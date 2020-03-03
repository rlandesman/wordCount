# TODO: Add stats for the word count found so user isnt just waiting with blank screenbbb
# TODO: huge block of text, takes user longer to read than code to runl.
# TODO: Cuztomizable output excel sheet name >>

import os
from openpyxl import Workbook
# from openpyxl.chart import (
#     ScatterChart,
#     Reference,
#     Series,
# )

## TODO: Read object from this file for abstraction
class DataEntry(object):
    publication_name = ""
    year = 0
    old = {} #i.e Traditional news reporting style
    new = {} #i.e Novel narrative-style news reporting
    newSum = 0
    oldSum = 0

    # The class "constructor" - It's actually an initializer
    def __init__(self, publication_name, year, old, new, oldSum, newSum):
        self.publication_name = publication_name
        self.year = year
        self.old = old
        self.new = new
        self.oldSum = oldSum
        self.newSum = newSum


def assignWordList(filename, thisDataEntry):
    """
    Go through the list of words, and seperate into old and new arrays
    Take DataEntry and assign appropriate dictionaries with word arrays, inti to 0 count
    static, no return
    """
    oldArr = []
    newArr = []
    try:
        with open(filename, encoding="latin-1") as file:
            lines = [line.rstrip() for line in file]
            idx = 0
            while(lines[idx] != "***"):
                oldArr.append(lines[idx].lower())
                idx += 1
            idx += 1 #Skip the delimitter
            for x in range(idx, len(lines)):
                newArr.append(lines[x].lower())
        file.close()
    except IOError:
        print("Error opening: " + str(filename))
    for x in oldArr:
        thisDataEntry.old[x] = 0
    for y in newArr:
        thisDataEntry.new[y] = 0

def parseFileName(filename):
    """
    Creates new DataEntry object and assigns appropriate fields according to filename
    Return the newly created DataEntry
    """
    entry = DataEntry("",0,{},{},0,0)
    wordArray = filename.split(".")
    entry.publication_name = wordArray[1]
    entry.year = wordArray[0]
    return entry

def getWordOccurance(filename,thisDataEntry):
    """
    Big boy function: Get the occurance of each key word in the articles
    Tweak dictionary in data entry, no return
    """
    try:
        with open(filename, encoding="latin-1") as file:
            cnt = -1
            lines = [line.rstrip() for line in file]
            for line in lines:
                lowerLine = line.lower()
                for key in thisDataEntry.old:
                    keyLength = len(key.split(" "))
                    if(keyLength == 1):
                        if key in lowerLine.split(" "): #By split checks if single word exists exactly
                            thisDataEntry.old[key] += 1
                    elif(keyLength > 1):
                        if key in lowerLine: #By split checks if single word exists exactly
                            thisDataEntry.old[key] += 1
                    else:
                        print("Error, key length unknown!")
                for key in thisDataEntry.new:
                    keyLength = len(key.split(" "))
                    if(keyLength == 1):
                        if key in lowerLine.split(" "):
                            thisDataEntry.new[key] += 1
                    elif(keyLength > 1):
                        if key in lowerLine:
                            thisDataEntry.new[key] += 1
                    else:
                        print("Error, key length unknown!")
        file.close()
    except IOError:
        print("Error opening: " + str(filename))

def performStats(dataArray):
    """
    Statically calculate and assign summed values of occurances to each entry
    """
    yearArray = [[0,0] for i in range(20)]
    for entry in dataArray:
        oSum = 0
        nSum = 0
        for k, v in entry.old.items():
            # print(k,v)
            oSum += v
        for k,v in entry.new.items():
            # print(k,v)
            nSum += v
        entry.oldSum = oSum
        entry.newSum = nSum
        idx = int(entry.year)%20 #0-19 index
        yearArray[idx][0] += entry.oldSum
        yearArray[idx][1] += entry.newSum
    return yearArray

def writeToExcel(statArray):
    wb = Workbook()
    ws = wb.active
    ws.title = "FrequencyChart"
    idx = nIdx = oIdx = 0
    flag = False
    for col in ws.iter_cols(min_row=1, min_col=2, max_col=21, max_row=1): #years instantiate
        if(idx < 10):
            yearString = "200" + str(idx)
        else:
            yearString = "20" + str(idx)
        for cell in col:
            cell.value = yearString
            idx += 1
    for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=1):
        for cell in row:
            if(not flag):
                cell.value = "Old"
                flag = True
            else:
                cell.value = "New"
    for col in ws.iter_rows(min_row=2, max_row=2, min_col=2, max_col=21):
        for cell in col:
            cell.value = statArray[oIdx][0]
            oIdx += 1
    for col in ws.iter_rows(min_row=3, max_row=3, min_col=2, max_col=21):
        for cell in col:
            cell.value = statArray[nIdx][1]
            nIdx += 1
    wb.save("outputWB.xlsx")

def main():
    directory = "/Users/roeelandesman/Desktop/wordCount/articles"
    word_list_file = "/Users/roeelandesman/Desktop/wordCount/word.list.txt"
    DataPointArray = []
    for filename in os.listdir(directory):
        if filename.endswith(".txt"):
            thisEntry = parseFileName(filename)
            assignWordList(word_list_file, thisEntry)
            fullPath = directory + "/" + filename
            getWordOccurance(fullPath, thisEntry)
            DataPointArray.append(thisEntry)

    allStatsArray = performStats(DataPointArray)
    writeToExcel(allStatsArray)

if __name__ == '__main__':
    main()
